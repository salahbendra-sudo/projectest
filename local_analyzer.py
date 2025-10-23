"""
Local Excel Analyzer - Basic analysis without AI dependency
Provides fallback functionality when OpenRouter API is unavailable
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from typing import Dict, List, Any, Optional
import json
import logging

logger = logging.getLogger(__name__)

class LocalExcelAnalyzer:
    """Basic Excel file analyzer for fallback functionality"""
    
    def __init__(self):
        self.supported_formulas = {
            'SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'IF', 'VLOOKUP', 
            'HLOOKUP', 'INDEX', 'MATCH', 'CONCATENATE', 'LEFT', 'RIGHT',
            'MID', 'LEN', 'FIND', 'SUBSTITUTE', 'ROUND', 'ROUNDUP', 
            'ROUNDDOWN', 'DATE', 'YEAR', 'MONTH', 'DAY', 'NOW', 'TODAY'
        }
    
    def analyze_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Analyze Excel file and return basic structure information
        """
        try:
            # Load workbook
            wb = load_workbook(file_path, read_only=True, data_only=False)
            
            analysis = {
                "file_info": {},
                "sheets": {},
                "formulas": [],
                "charts": [],
                "structure": {},
                "analysis_summary": {}
            }
            
            # File information
            analysis["file_info"] = {
                "filename": file_path.split('/')[-1],
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames
            }
            
            # Analyze each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_analysis = self._analyze_sheet(ws, sheet_name)
                analysis["sheets"][sheet_name] = sheet_analysis
            
            # Generate analysis summary
            analysis["analysis_summary"] = self._generate_summary(analysis)
            
            wb.close()
            return analysis
            
        except Exception as e:
            logger.error(f"Error analyzing Excel file: {e}")
            return {
                "error": str(e),
                "file_info": {"filename": file_path.split('/')[-1]},
                "sheets": {},
                "formulas": [],
                "charts": [],
                "structure": {},
                "analysis_summary": {"error": "Analysis failed"}
            }
    
    def _analyze_sheet(self, worksheet, sheet_name: str) -> Dict[str, Any]:
        """Analyze individual worksheet"""
        sheet_data = {
            "name": sheet_name,
            "dimensions": f"{worksheet.max_row} rows x {worksheet.max_column} columns",
            "data_range": f"A1:{openpyxl.utils.get_column_letter(worksheet.max_column)}{worksheet.max_row}",
            "formula_count": 0,
            "data_types": {},
            "sample_data": {}
        }
        
        # Get sample data (first 5 rows)
        sample_rows = []
        for i, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if i > 5:  # Limit to first 5 rows
                break
            sample_rows.append(list(row) if row else [])
        
        sheet_data["sample_data"] = sample_rows
        
        # Count formulas
        formula_count = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    formula_count += 1
        
        sheet_data["formula_count"] = formula_count
        
        return sheet_data
    
    def _generate_summary(self, analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Generate summary of the analysis"""
        sheets = analysis["sheets"]
        total_formulas = sum(sheet.get("formula_count", 0) for sheet in sheets.values())
        
        # Determine template type based on structure
        template_type = self._detect_template_type(analysis)
        
        return {
            "total_sheets": len(sheets),
            "total_formulas": total_formulas,
            "template_type": template_type,
            "complexity": self._assess_complexity(total_formulas, len(sheets)),
            "recommended_framework": "Streamlit",
            "estimated_development_time": f"{max(1, total_formulas // 10)} hours"
        }
    
    def _detect_template_type(self, analysis: Dict[str, Any]) -> str:
        """Detect the type of template based on sheet structure"""
        sheets = analysis["sheets"]
        
        if not sheets:
            return "Unknown"
        
        sheet_names = [name.lower() for name in sheets.keys()]
        
        # Check for common template patterns
        financial_keywords = ['financial', 'revenue', 'profit', 'income', 'balance', 'cash flow']
        sales_keywords = ['sales', 'revenue', 'customer', 'product', 'region']
        inventory_keywords = ['inventory', 'stock', 'supplier', 'order']
        
        for sheet_name in sheet_names:
            if any(keyword in sheet_name for keyword in financial_keywords):
                return "Financial Model"
            elif any(keyword in sheet_name for keyword in sales_keywords):
                return "Sales Dashboard"
            elif any(keyword in sheet_name for keyword in inventory_keywords):
                return "Inventory Management"
        
        return "Data Analysis"
    
    def _assess_complexity(self, total_formulas: int, total_sheets: int) -> str:
        """Assess complexity of the Excel file"""
        if total_sheets <= 1 and total_formulas <= 5:
            return "Simple"
        elif total_sheets <= 3 and total_formulas <= 20:
            return "Medium"
        elif total_sheets <= 5 and total_formulas <= 50:
            return "Complex"
        else:
            return "Very Complex"
    
    def generate_basic_app_code(self, analysis: Dict[str, Any]) -> str:
        """Generate basic Streamlit app code based on analysis"""
        
        template = '''
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

# Page configuration
st.set_page_config(
    page_title="Excel Data Explorer",
    page_icon="📊",
    layout="wide"
)

# Custom CSS
st.markdown("""
<style>
    .main > div {
        padding-top: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# Title
st.title("📊 Excel Data Explorer")
st.markdown("Generated from your Excel file")

# File upload
uploaded_file = st.file_uploader("Upload Excel File", type=['xlsx', 'xls'])

if uploaded_file is not None:
    try:
        # Load data
        sheets = pd.read_excel(uploaded_file, sheet_name=None)
        
        # Create tabs for each sheet
        tabs = st.tabs(list(sheets.keys()))
        
        for i, (sheet_name, df) in enumerate(sheets.items()):
            with tabs[i]:
                st.header(f"Sheet: {sheet_name}")
                
                # Data overview
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Rows", len(df))
                with col2:
                    st.metric("Columns", len(df.columns))
                with col3:
                    st.metric("Missing Values", df.isnull().sum().sum())
                
                # Data preview
                st.subheader("Data Preview")
                st.dataframe(df, use_container_width=True)
                
                # Basic statistics
                st.subheader("Statistics")
                st.dataframe(df.describe(), use_container_width=True)
                
                # Simple visualizations for numeric columns
                numeric_cols = df.select_dtypes(include=['number']).columns
                if len(numeric_cols) > 0:
                    st.subheader("Visualizations")
                    
                    # Column selection for charts
                    selected_col = st.selectbox(
                        "Select column for visualization",
                        numeric_cols,
                        key=f"col_select_{sheet_name}"
                    )
                    
                    if selected_col:
                        # Histogram
                        fig = px.histogram(df, x=selected_col, title=f"Distribution of {selected_col}")
                        st.plotly_chart(fig, use_container_width=True)
        
        # Download processed data
        st.sidebar.header("Data Export")
        for sheet_name, df in sheets.items():
            csv = df.to_csv(index=False)
            st.sidebar.download_button(
                label=f"Download {sheet_name} as CSV",
                data=csv,
                file_name=f"{sheet_name}_data.csv",
                mime="text/csv"
            )
                
    except Exception as e:
        st.error(f"Error processing file: {e}")
else:
    st.info("Please upload an Excel file to begin")

st.sidebar.markdown("---")
st.sidebar.info("This app was automatically generated from your Excel file")
'''
        
        return template

# Utility function for quick analysis
def quick_analyze(file_path: str) -> Dict[str, Any]:
    """Quick analysis of Excel file"""
    analyzer = LocalExcelAnalyzer()
    return analyzer.analyze_excel_file(file_path)

if __name__ == "__main__":
    # Test the analyzer
    import sys
    if len(sys.argv) > 1:
        result = quick_analyze(sys.argv[1])
        print(json.dumps(result, indent=2))
    else:
        print("Usage: python local_analyzer.py <excel_file_path>")