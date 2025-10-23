"""
Enhanced Excel Analyzer
Comprehensive analysis of Excel files including formulas, charts, business logic, and structure
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from typing import Dict, List, Any, Optional, Tuple
import json
import logging
import re
from pathlib import Path
import io
import tempfile
import os

logger = logging.getLogger(__name__)

class EnhancedExcelAnalyzer:
    """Comprehensive Excel file analyzer with deep structure and logic extraction"""
    
    def __init__(self):
        self.supported_formulas = {
            'SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'IF', 'VLOOKUP', 'HLOOKUP',
            'INDEX', 'MATCH', 'CONCATENATE', 'LEFT', 'RIGHT', 'MID', 'LEN', 'FIND',
            'SUBSTITUTE', 'ROUND', 'ROUNDUP', 'ROUNDDOWN', 'DATE', 'YEAR', 'MONTH',
            'DAY', 'NOW', 'TODAY', 'PMT', 'FV', 'PV', 'IRR', 'NPV', 'RATE',
            'SLN', 'DDB', 'VDB', 'SUMIF', 'SUMIFS', 'COUNTIF', 'COUNTIFS',
            'AVERAGEIF', 'AVERAGEIFS', 'AND', 'OR', 'NOT', 'XOR'
        }
        
        self.business_patterns = {
            'financial': ['revenue', 'profit', 'income', 'expense', 'cost', 'margin', 'cash', 'flow', 'balance',
                         'cogs', 'gross', 'operating', 'net', 'financial', 'budget', 'tax', 'rate'],
            'sales': ['sales', 'customer', 'product', 'region', 'territory', 'quota', 'commission', 'units'],
            'inventory': ['inventory', 'stock', 'supplier', 'order', 'quantity', 'reorder', 'warehouse', 'stock'],
            'hr': ['employee', 'salary', 'department', 'position', 'hire', 'termination', 'benefits', 'hr'],
            'project': ['project', 'task', 'milestone', 'deadline', 'resource', 'budget', 'timeline']
        }
    
    def analyze_excel_comprehensively(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Perform comprehensive analysis of Excel file
        """
        try:
            file_ext = Path(filename).suffix.lower()
            
            # Handle CSV files differently
            if file_ext == '.csv':
                return self._analyze_csv_file(file_content, filename)
            
            # Save file temporarily for analysis
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                tmp_file.write(file_content)
                tmp_path = tmp_file.name
            
            # Load workbook
            wb = load_workbook(tmp_path, read_only=True, data_only=False)
            
            analysis = {
                'file_info': {},
                'structure': {},
                'formulas': {},
                'charts': {},
                'business_logic': {},
                'data_relationships': {},
                'user_interaction': {},
                'analysis_summary': {}
            }
            
            # File information
            analysis['file_info'] = self._analyze_file_info(filename, wb)
            
            # Structure analysis
            analysis['structure'] = self._analyze_structure(wb)
            
            # Formula analysis
            analysis['formulas'] = self._analyze_formulas(wb)
            
            # Chart analysis
            analysis['charts'] = self._analyze_charts(wb)
            
            # Business logic inference
            analysis['business_logic'] = self._infer_business_logic(wb, analysis)
            
            # Data relationships
            analysis['data_relationships'] = self._analyze_relationships(wb)
            
            # User interaction patterns
            analysis['user_interaction'] = self._analyze_user_interaction(wb)
            
            # Generate comprehensive summary
            analysis['analysis_summary'] = self._generate_comprehensive_summary(analysis)
            
            wb.close()
            
            # Cleanup
            os.unlink(tmp_path)
            
            return analysis
            
        except Exception as e:
            logger.error(f"Comprehensive analysis failed: {e}")
            return {
                'error': str(e),
                'file_info': {'filename': filename},
                'analysis_summary': {'error': 'Comprehensive analysis failed'}
            }
    
    def _analyze_file_info(self, filename: str, wb) -> Dict[str, Any]:
        """Analyze basic file information"""
        return {
            'filename': filename,
            'sheet_count': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
            'file_size': len(filename)  # Approximate
        }
    
    def _analyze_structure(self, wb) -> Dict[str, Any]:
        """Analyze workbook structure"""
        structure = {
            'sheets': {},
            'total_rows': 0,
            'total_columns': 0,
            'data_density': 0.0
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_analysis = self._analyze_sheet_structure(ws, sheet_name)
            structure['sheets'][sheet_name] = sheet_analysis
            structure['total_rows'] += sheet_analysis['rows']
            structure['total_columns'] = max(structure['total_columns'], sheet_analysis['columns'])
        
        return structure
    
    def _analyze_sheet_structure(self, ws, sheet_name: str) -> Dict[str, Any]:
        """Analyze individual sheet structure"""
        # Get dimensions
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        
        # Analyze data types and patterns
        data_types = {}
        sample_data = []
        headers = []
        
        # Get first few rows for analysis
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > 10:  # Limit to first 10 rows
                break
            if row:
                if i == 1:
                    # First row is headers
                    headers = [str(cell) if cell is not None else f"Column_{j+1}" for j, cell in enumerate(row)]
                    # Initialize data_types with headers
                    for header in headers:
                        data_types[header] = set()
                else:
                    sample_data.append(list(row))
                    
                    # Analyze data types in this row
                    for j, cell in enumerate(row):
                        if cell is not None and j < len(headers):
                            col_name = headers[j]
                            data_types[col_name].add(type(cell).__name__)
        
        # Check for charts (simplified)
        has_charts = False
        if hasattr(ws, '_charts'):
            try:
                has_charts = len(ws._charts) > 0
            except:
                has_charts = False
        
        return {
            'name': sheet_name,
            'rows': max_row,
            'columns': max_col,
            'data_range': f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}" if max_row > 0 and max_col > 0 else "Empty",
            'data_types': {k: list(v) for k, v in data_types.items()},
            'sample_data': sample_data,
            'has_formulas': any(cell.data_type == 'f' for row in ws.iter_rows() for cell in row),
            'has_charts': has_charts
        }
    
    def _analyze_formulas(self, wb) -> Dict[str, Any]:
        """Extract and analyze formulas"""
        formulas = {
            'total_count': 0,
            'by_sheet': {},
            'complexity_analysis': {},
            'formula_categories': {}
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_formulas = self._extract_sheet_formulas(ws, sheet_name)
            formulas['by_sheet'][sheet_name] = sheet_formulas
            formulas['total_count'] += sheet_formulas['count']
        
        # Analyze formula complexity
        formulas['complexity_analysis'] = self._analyze_formula_complexity(formulas)
        formulas['formula_categories'] = self._categorize_formulas(formulas)
        
        return formulas
    
    def _extract_sheet_formulas(self, ws, sheet_name: str) -> Dict[str, Any]:
        """Extract formulas from a single sheet"""
        formulas = []
        formula_count = 0
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    formula_count += 1
                    formula_info = {
                        'cell': cell.coordinate,
                        'formula': str(cell.value),
                        'category': self._categorize_formula(str(cell.value))
                    }
                    formulas.append(formula_info)
        
        return {
            'count': formula_count,
            'formulas': formulas,
            'complexity': 'High' if formula_count > 20 else 'Medium' if formula_count > 5 else 'Low'
        }
    
    def _categorize_formula(self, formula: str) -> str:
        """Categorize formula by type"""
        formula_upper = formula.upper()
        
        if any(f"{func}(" in formula_upper for func in ['SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN']):
            return 'Aggregation'
        elif any(f"{func}(" in formula_upper for func in ['IF', 'AND', 'OR', 'NOT']):
            return 'Logical'
        elif any(f"{func}(" in formula_upper for func in ['VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH']):
            return 'Lookup'
        elif any(f"{func}(" in formula_upper for func in ['PMT', 'FV', 'PV', 'IRR', 'NPV']):
            return 'Financial'
        elif any(f"{func}(" in formula_upper for func in ['DATE', 'YEAR', 'MONTH', 'DAY']):
            return 'Date/Time'
        elif any(f"{func}(" in formula_upper for func in ['LEFT', 'RIGHT', 'MID', 'LEN', 'FIND']):
            return 'Text'
        else:
            return 'Other'
    
    def _analyze_formula_complexity(self, formulas: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze formula complexity"""
        total_formulas = formulas['total_count']
        
        # Count formula types across all sheets
        formula_categories = {}
        for sheet_name, sheet_data in formulas['by_sheet'].items():
            for formula_info in sheet_data['formulas']:
                category = formula_info['category']
                formula_categories[category] = formula_categories.get(category, 0) + 1
        
        return {
            'total_formulas': total_formulas,
            'category_distribution': formula_categories,
            'complexity_level': 'High' if total_formulas > 50 else 'Medium' if total_formulas > 10 else 'Low'
        }
    
    def _categorize_formulas(self, formulas: Dict[str, Any]) -> Dict[str, Any]:
        """Categorize formulas by business function"""
        categories = {}
        
        for sheet_name, sheet_data in formulas['by_sheet'].items():
            for formula_info in sheet_data['formulas']:
                category = formula_info['category']
                if category not in categories:
                    categories[category] = []
                categories[category].append({
                    'sheet': sheet_name,
                    'cell': formula_info['cell'],
                    'formula': formula_info['formula']
                })
        
        return categories
    
    def _analyze_csv_file(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Analyze CSV file"""
        try:
            # Read CSV data
            df = pd.read_csv(io.BytesIO(file_content))
            
            analysis = {
                'file_info': {
                    'filename': filename,
                    'sheet_count': 1,
                    'sheet_names': ['CSV Data'],
                    'file_type': '.csv'
                },
                'structure': {
                    'sheets': {
                        'CSV Data': {
                            'name': 'CSV Data',
                            'rows': len(df),
                            'columns': len(df.columns),
                            'data_range': f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}{len(df)}",
                            'data_types': {col: [str(df[col].dtype)] for col in df.columns},
                            'sample_data': df.head(5).values.tolist(),
                            'has_formulas': False,
                            'has_charts': False
                        }
                    },
                    'total_rows': len(df),
                    'total_columns': len(df.columns),
                    'data_density': 1.0
                },
                'formulas': {
                    'total_count': 0,
                    'by_sheet': {},
                    'complexity_analysis': {},
                    'formula_categories': {}
                },
                'charts': {
                    'total_count': 0,
                    'by_sheet': {},
                    'chart_types': {}
                },
                'business_logic': {
                    'domain': self._detect_csv_domain(df),
                    'key_operations': [],
                    'calculations': [],
                    'workflow_patterns': [],
                    'business_rules': []
                },
                'data_relationships': {
                    'cross_sheet_references': [],
                    'data_flow': [],
                    'dependencies': []
                },
                'user_interaction': {
                    'input_cells': [],
                    'output_cells': [],
                    'interactive_elements': []
                },
                'analysis_summary': {
                    'overall_complexity': 'Low',
                    'recommended_template': 'Universal Data Explorer',
                    'transformation_confidence': 0.9,
                    'key_findings': [f"CSV file with {len(df)} rows and {len(df.columns)} columns"],
                    'recommendations': ["Use universal data explorer for CSV files"]
                }
            }
            
            return analysis
            
        except Exception as e:
            logger.error(f"CSV analysis failed: {e}")
            return {
                'error': str(e),
                'file_info': {'filename': filename},
                'analysis_summary': {'error': 'CSV analysis failed'}
            }
    
    def _detect_csv_domain(self, df: pd.DataFrame) -> str:
        """Detect business domain from CSV data"""
        column_names = [col.lower() for col in df.columns]
        
        for domain, keywords in self.business_patterns.items():
            for col_name in column_names:
                if any(keyword in col_name for keyword in keywords):
                    return domain.capitalize()
        
        return 'General Data Analysis'
    
    def _analyze_charts(self, wb) -> Dict[str, Any]:
        """Analyze charts in the workbook"""
        charts = {
            'total_count': 0,
            'by_sheet': {},
            'chart_types': {}
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Try to detect charts (simplified approach)
            # In read-only mode, we can't access _charts directly
            # This is a simplified detection
            chart_count = 0
            
            # Check for common chart indicators
            # This is a basic implementation - in production, we'd use a more robust approach
            if hasattr(ws, '_charts'):
                try:
                    chart_count = len(ws._charts)
                except:
                    chart_count = 0
            
            charts['total_count'] += chart_count
            
            if chart_count > 0:
                charts['by_sheet'][sheet_name] = {
                    'count': chart_count,
                    'charts': [f"Chart_{i+1}" for i in range(chart_count)]
                }
        
        return charts
    
    def _infer_business_logic(self, wb, analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Infer business logic from structure and formulas"""
        business_logic = {
            'domain': 'Unknown',
            'key_operations': [],
            'calculations': [],
            'workflow_patterns': [],
            'business_rules': []
        }
        
        # Determine domain based on sheet names and content
        business_logic['domain'] = self._detect_business_domain(analysis)
        
        # Extract key operations from formulas
        business_logic['key_operations'] = self._extract_key_operations(analysis['formulas'])
        
        # Identify workflow patterns
        business_logic['workflow_patterns'] = self._identify_workflow_patterns(analysis['structure'])
        
        # Extract business rules
        business_logic['business_rules'] = self._extract_business_rules(analysis['formulas'])
        
        return business_logic
    
    def _detect_business_domain(self, analysis: Dict[str, Any]) -> str:
        """Detect the business domain based on sheet names and content"""
        sheet_names = [name.lower() for name in analysis['structure']['sheets'].keys()]
        
        # Check sheet names first
        for domain, keywords in self.business_patterns.items():
            for sheet_name in sheet_names:
                if any(keyword in sheet_name for keyword in keywords):
                    return domain.capitalize()
        
        # Check column names across all sheets
        domain_scores = {domain: 0 for domain in self.business_patterns.keys()}
        
        for sheet_name, sheet_data in analysis['structure']['sheets'].items():
            # Get column names from data types
            if 'data_types' in sheet_data:
                column_names = list(sheet_data['data_types'].keys())
                for col_name in column_names:
                    col_lower = str(col_name).lower()
                    for domain, keywords in self.business_patterns.items():
                        if any(keyword in col_lower for keyword in keywords):
                            domain_scores[domain] += 2
        
        # Find domain with highest score from column names
        if domain_scores:
            max_domain = max(domain_scores, key=domain_scores.get)
            if domain_scores[max_domain] >= 2:  # Require at least 2 matches
                return max_domain.capitalize()
        
        # Check formula categories for hints
        formula_categories = analysis['formulas'].get('formula_categories', {})
        if 'Financial' in formula_categories:
            return 'Financial'
        elif 'Lookup' in formula_categories and len(formula_categories.get('Lookup', [])) > 5:
            return 'Inventory'
        
        return 'General Data Analysis'
    
    def _extract_key_operations(self, formulas: Dict[str, Any]) -> List[str]:
        """Extract key business operations from formulas"""
        operations = []
        
        for sheet_name, sheet_data in formulas['by_sheet'].items():
            for formula_info in sheet_data['formulas']:
                formula = formula_info['formula']
                category = formula_info['category']
                
                if category == 'Financial':
                    operations.append(f"Financial calculation in {sheet_name}: {formula}")
                elif category == 'Aggregation' and 'SUM' in formula.upper():
                    operations.append(f"Summation in {sheet_name}")
                elif category == 'Lookup':
                    operations.append(f"Data lookup in {sheet_name}")
        
        return operations
    
    def _identify_workflow_patterns(self, structure: Dict[str, Any]) -> List[str]:
        """Identify workflow patterns from sheet structure"""
        patterns = []
        sheet_names = list(structure['sheets'].keys())
        
        if len(sheet_names) > 1:
            # Check for common workflow patterns
            if any('input' in name.lower() for name in sheet_names) and any('output' in name.lower() for name in sheet_names):
                patterns.append("Input → Processing → Output workflow")
            
            if any('raw' in name.lower() for name in sheet_names) and any('processed' in name.lower() for name in sheet_names):
                patterns.append("Data transformation workflow")
            
            if any('summary' in name.lower() for name in sheet_names):
                patterns.append("Detailed data → Summary workflow")
        
        return patterns
    
    def _extract_business_rules(self, formulas: Dict[str, Any]) -> List[str]:
        """Extract business rules from logical formulas"""
        rules = []
        
        for sheet_name, sheet_data in formulas['by_sheet'].items():
            for formula_info in sheet_data['formulas']:
                if formula_info['category'] == 'Logical':
                    formula = formula_info['formula']
                    # Extract simple IF statements as business rules
                    if 'IF(' in formula.upper():
                        rules.append(f"Business rule in {sheet_name}: {formula}")
        
        return rules
    
    def _analyze_relationships(self, wb) -> Dict[str, Any]:
        """Analyze data relationships between sheets"""
        relationships = {
            'cross_sheet_references': [],
            'data_flow': [],
            'dependencies': []
        }
        
        # This is a simplified analysis - in a real implementation,
        # we would analyze formula references between sheets
        sheet_names = wb.sheetnames
        
        if len(sheet_names) > 1:
            relationships['cross_sheet_references'].append(
                f"Multiple sheets ({len(sheet_names)}) suggest data relationships"
            )
        
        return relationships
    
    def _analyze_user_interaction(self, wb) -> Dict[str, Any]:
        """Analyze user interaction patterns"""
        interaction = {
            'input_cells': [],
            'output_cells': [],
            'interactive_elements': []
        }
        
        # Simplified analysis - in real implementation, we would identify
        # cells that are likely inputs (no formulas, manual entry)
        # and outputs (formula results)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Check for data validation (indicates user input)
            if hasattr(ws, 'data_validations') and ws.data_validations:
                interaction['input_cells'].append(f"Data validation in {sheet_name}")
            
            # Check for conditional formatting (indicates visual feedback)
            if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
                interaction['interactive_elements'].append(f"Conditional formatting in {sheet_name}")
        
        return interaction
    
    def _generate_comprehensive_summary(self, analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Generate comprehensive analysis summary"""
        summary = {
            'overall_complexity': 'Medium',
            'recommended_template': 'Universal Data Explorer',
            'transformation_confidence': 0.8,
            'key_findings': [],
            'recommendations': []
        }
        
        # Assess complexity
        total_formulas = analysis['formulas']['total_count']
        total_sheets = analysis['structure']['total_rows']
        
        if total_formulas > 50 or total_sheets > 1000:
            summary['overall_complexity'] = 'High'
        elif total_formulas < 5 and total_sheets < 100:
            summary['overall_complexity'] = 'Low'
        
        # Determine recommended template
        domain = analysis['business_logic']['domain']
        if domain == 'Financial':
            summary['recommended_template'] = 'Financial Dashboard'
        elif domain == 'Sales':
            summary['recommended_template'] = 'Sales Analytics'
        elif domain == 'Inventory':
            summary['recommended_template'] = 'Inventory Management'
        elif domain == 'Hr':
            summary['recommended_template'] = 'HR Management'
        elif domain == 'Project':
            summary['recommended_template'] = 'Project Management'
        
        # Generate key findings
        if analysis['formulas']['total_count'] > 0:
            summary['key_findings'].append(f"Found {analysis['formulas']['total_count']} formulas")
        
        if analysis['charts']['total_count'] > 0:
            summary['key_findings'].append(f"Found {analysis['charts']['total_count']} charts")
        
        if analysis['business_logic']['domain'] != 'General Data Analysis':
            summary['key_findings'].append(f"Detected {analysis['business_logic']['domain']} domain")
        
        # Generate recommendations
        if analysis['formulas']['total_count'] > 20:
            summary['recommendations'].append("Consider implementing formula calculations in Python")
        
        if analysis['charts']['total_count'] > 0:
            summary['recommendations'].append("Recreate charts using Plotly for interactivity")
        
        if len(analysis['business_logic']['business_rules']) > 0:
            summary['recommendations'].append("Implement business rules as validation logic")
        
        return summary

# Utility function for quick analysis
def quick_comprehensive_analysis(file_content: bytes, filename: str) -> Dict[str, Any]:
    """Quick comprehensive analysis of Excel file"""
    analyzer = EnhancedExcelAnalyzer()
    return analyzer.analyze_excel_comprehensively(file_content, filename)

if __name__ == "__main__":
    # Test the analyzer
    import sys
    if len(sys.argv) > 1:
        with open(sys.argv[1], 'rb') as f:
            file_content = f.read()
        result = quick_comprehensive_analysis(file_content, sys.argv[1])
        print(json.dumps(result, indent=2))
    else:
        print("Usage: python enhanced_excel_analyzer.py <excel_file_path>")