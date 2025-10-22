import os
import re
import tempfile
import uuid
import zipfile
import logging
from fastapi import FastAPI, UploadFile, Form, File, HTTPException
from fastapi.responses import StreamingResponse
import pandas as pd
from openai import OpenAI
from pathlib import Path
from typing import Optional, Dict, Any, List, Set, Tuple
import shutil
from dotenv import load_dotenv
import io
import asyncio
import openpyxl
from openpyxl import load_workbook
import xlrd
import oletools.olevba as olevba
import requests
import json
from datetime import datetime
import traceback

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

load_dotenv()

app = FastAPI(
    title="Excel-to-Project Generator API",
    description="Analyzes Excel files and generates Python projects based on the analysis",
    version="3.1.0",
    docs_url="/docs",
    redoc_url=None
)

# Configuration
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
if not OPENROUTER_API_KEY:
    raise ValueError("Missing OPENROUTER_API_KEY")

OPENROUTER_API_URL = "https://openrouter.ai/api/v1/chat/completions"
MAX_EXCEL_SIZE = 10 * 1024 * 1024  # 10MB
MAX_PROJECT_SIZE = 5 * 1024 * 1024  # 5MB

# --- Helper Functions ---
def get_openai_client():
    """Initialize OpenRouter client"""
    return OpenAI(
        api_key=OPENROUTER_API_KEY,
        base_url="https://openrouter.ai/api/v1"
    )

def _get_column_letter(col_idx: int) -> str:
    """Convert column index to Excel letter"""
    string = ""
    while col_idx >= 0:
        string = chr(col_idx % 26 + ord('A')) + string
        col_idx = col_idx // 26 - 1
    return string

def extract_vba_code(file_path: Path) -> Optional[str]:
    """Extract VBA code using oletools"""
    try:
        vba_parser = olevba.VBA_Parser(str(file_path))
        if not vba_parser.detect_vba_macros():
            return None
        
        vba_code = []
        for (filename, stream_path, vba_filename, code) in vba_parser.extract_macros():
            if code and code.strip():
                vba_code.append(f"=== Module: {vba_filename} in {stream_path} ===\n{code}\n")
        
        vba_parser.close()
        return "\n".join(vba_code) if vba_code else None
    except Exception as e:
        logger.warning(f"VBA extraction failed: {str(e)}")
        return None

def extract_formulas_xlsx(file_path: Path) -> Dict[str, Dict[str, str]]:
    """Extract formulas from .xlsx files"""
    formulas = {}
    try:
        wb = load_workbook(str(file_path), data_only=False, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_formulas = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':
                        formula_value = str(cell.value)
                        if not formula_value.startswith('='):
                            formula_value = f"={formula_value}"
                        sheet_formulas[cell.coordinate] = formula_value
            if sheet_formulas:
                formulas[sheet_name] = sheet_formulas
        return formulas
    except Exception as e:
        logger.error(f"XLSX formula extraction failed: {str(e)}")
        return {}

def extract_formulas_xls(file_path: Path) -> Dict[str, Dict[str, str]]:
    """Extract formulas from .xls files"""
    formulas = {}
    try:
        workbook = xlrd.open_workbook(str(file_path), on_demand=True)
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            sheet_formulas = {}
            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    if cell.ctype == xlrd.XL_CELL_FORMULA:
                        col_letter = _get_column_letter(col_idx)
                        cell_coord = f"{col_letter}{row_idx + 1}"
                        sheet_formulas[cell_coord] = f"={cell.value}"
            if sheet_formulas:
                formulas[sheet_name] = sheet_formulas
        return formulas
    except Exception as e:
        logger.error(f"XLS formula extraction failed: {str(e)}")
        return {}

def extract_formulas(file_path: Path) -> Dict[str, Dict[str, str]]:
    """Extract formulas based on file type"""
    suffix = file_path.suffix.lower()
    if suffix in ['.xlsx', '.xlsm']:
        return extract_formulas_xlsx(file_path)
    elif suffix == '.xls':
        return extract_formulas_xls(file_path)
    return {}

def get_formula_dependencies(formulas: Dict[str, Dict[str, str]]) -> Dict[str, Set[str]]:
    """Identify cells referenced by formulas"""
    deps = {}
    cell_ref_pattern = re.compile(r'[A-Za-z]{1,3}\d{1,7}')
    for sheet, sheet_formulas in formulas.items():
        sheet_deps = set()
        for formula in sheet_formulas.values():
            search_part = formula.split('!')[-1]
            matches = cell_ref_pattern.findall(search_part)
            sheet_deps.update(m.upper() for m in matches)
        deps[sheet] = sheet_deps
    return deps

def select_rows(df: pd.DataFrame, formula_deps: Set[str]) -> pd.DataFrame:
    """Smart row selection with formula-aware sampling"""
    if df.empty:
        return df

    selected_indices = set()
    num_rows = len(df)
    
    # Always include key rows
    key_indices = {0, num_rows // 2, num_rows - 1}
    if num_rows > 1:
        key_indices.add(1)
    if num_rows > 2:
        key_indices.add(num_rows - 2)
    selected_indices.update(idx for idx in key_indices if 0 <= idx < num_rows)

    # Add formula dependencies
    for dep in formula_deps:
        if match := re.search(r'\d+', dep):
            row_num = int(match.group(0)) - 1
            if 0 <= row_num < num_rows:
                selected_indices.add(row_num)
    
    # Fill with additional rows if needed
    for idx in df.index:
        if len(selected_indices) >= 50:
            break
        if idx not in selected_indices and not df.loc[idx].isnull().all():
            selected_indices.add(idx)

    return df.loc[sorted(selected_indices)].copy()

def generate_analysis_prompt(data: Dict[str, pd.DataFrame], formulas: Dict, vba: Optional[str], instructions: str) -> str:
    """Generate LLM prompt for analysis"""
    data_csv = "\n".join(
        f"\n\n**Sheet: {sheet}**\n```csv\n{df.to_csv(index=False)}\n```"
        for sheet, df in data.items() if not df.empty
    ) or "\n\nNo data rows sampled or all sheets are empty."

    formulas_str = "\n".join(
        f"\n\n**Sheet: {sheet}**\n" + "\n".join(f"- {cell}: {formula}" for cell, formula in list(formulas.items())[:100])
        for sheet, formulas in formulas.items() if formulas
    ) if formulas else "\n\nNo formulas detected."

    vba_section = f"\n\n**VBA CODE:**\n```vb\n{vba[:5000]}\n```" if vba else "\n\nNo VBA code detected."

    return f"""**INSTRUCTIONS:**
{instructions}

**DATA SAMPLES:**{data_csv}

**FORMULAS:**{formulas_str}

**VBA CODE:**{vba_section}

**ANALYSIS GUIDANCE:**
1. Identify key data structures and business logic
2. Note calculations that should become API endpoints
3. Flag potential security issues in VBA
4. Suggest database models based on data
5. Highlight any data validation rules
"""

def create_generation_prompt(analysis: str, project_prompt: str) -> str:
    """Generate project creation prompt with strict formatting"""
    return f"""**PROJECT REQUIREMENTS:**
{project_prompt}

**ANALYSIS CONTEXT:**
{analysis}

**YOU MUST FOLLOW THESE EXACT INSTRUCTIONS:**

1. Generate ALL required project files using THIS EXACT FORMAT for EACH file:

FILENAME: relative/path/to/file.ext
CONTENT:
[EXACT FILE CONTENT STARTS HERE]
... file content ...
[EXACT FILE CONTENT ENDS HERE]
====

2. Requirements:
- First file MUST be main.py or app.py
- Include ALL necessary files (Python, configs, etc.)
- Use PROPER SYNTAX for each file type
- MAINTAIN EXACT INDENTATION
- Include ALL imports and dependencies
- Preserve Excel business logic

3. Example VALID response:
FILENAME: main.py
CONTENT:
from fastapi import FastAPI

app = FastAPI()

@app.get("/")
def read_root():
    return {{"message": "Hello World"}}
====
FILENAME: requirements.txt
CONTENT:
fastapi>=0.68.0
uvicorn>=0.15.0
====
"""

async def call_llm(prompt: str, model: str, temperature: float = 0.3) -> str:
    """Call LLM with error handling"""
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": max(0.1, min(temperature, 1.0))
    }
    
    try:
        response = requests.post(OPENROUTER_API_URL, headers=headers, json=payload)
        response.raise_for_status()
        return response.json()["choices"][0]["message"]["content"]
    except Exception as e:
        logger.error(f"API call failed: {str(e)}")
        raise HTTPException(502, f"AI service error: {str(e)}")

def parse_generated_files(llm_response: str) -> Dict[str, str]:
    """Parse LLM response into files with enhanced validation"""
    files = {}
    current_file = None
    current_content = []
    in_content = False
    found_files = False
    
    for line in llm_response.splitlines():
        stripped = line.strip()
        
        if stripped.startswith("FILENAME:"):
            found_files = True
            if current_file and current_content:
                files[current_file] = "\n".join(current_content).strip()
            current_file = stripped[len("FILENAME:"):].strip()
            if not current_file:  # Skip empty filenames
                current_file = None
                continue
            current_content = []
            in_content = False
        
        elif stripped == "CONTENT:":
            if current_file:  # Only start content if we have a filename
                in_content = True
        
        elif stripped == "====":
            if current_file and current_content:
                files[current_file] = "\n".join(current_content).strip()
            current_file = None
            current_content = []
            in_content = False
        
        elif in_content and current_file:
            current_content.append(line)
    
    # Add the last file if exists
    if current_file and current_content:
        files[current_file] = "\n".join(current_content).strip()
    
    if not found_files:
        logger.error(f"No valid file markers found in response. Response start:\n{llm_response[:500]}")
    
    return files

def validate_python_syntax(content: str) -> bool:
    """Basic Python syntax check"""
    try:
        compile(content, "<string>", "exec")
        return True
    except SyntaxError:
        return False

def validate_generated_files(files: Dict[str, str]) -> Tuple[bool, List[str]]:
    """Validate generated files and return (is_valid, errors)"""
    errors = []
    
    if not files:
        errors.append("No files were generated")
        return (False, errors)
    
    # Check for at least one Python file
    python_files = [f for f in files if f.endswith('.py')]
    if not python_files:
        errors.append("No Python files were generated")
    
    # Validate Python syntax
    for filename, content in files.items():
        if filename.endswith('.py'):
            if not validate_python_syntax(content):
                errors.append(f"Invalid Python syntax in {filename}")
    
    return (len(errors) == 0, errors)

async def process_excel(file_path: Path) -> Dict[str, Any]:
    """Process Excel file into structured data"""
    result = {
        "data": {},
        "formulas": extract_formulas(file_path),
        "vba": extract_vba_code(file_path),
        "metadata": {}
    }
    
    try:
        all_sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
        formula_deps = {
            sheet: set(re.findall(r'[A-Z]+\d+', " ".join(formulas.values())))
            for sheet, formulas in result["formulas"].items()
        }
        
        for sheet, df in all_sheets.items():
            result["data"][sheet] = select_rows(df, formula_deps.get(sheet, set()))
        
        result["metadata"]["sheets"] = len(all_sheets)
        result["metadata"]["rows"] = sum(len(df) for df in result["data"].values())
        return result
    except Exception as e:
        logger.error(f"Excel processing failed: {str(e)}")
        raise ValueError(f"Excel processing error: {str(e)}")

async def generate_zip_output(analysis: str, generated_files: Dict[str, str], original_excel: Path) -> StreamingResponse:
    """Create final ZIP output with validation"""
    # Validate files before creating ZIP
    is_valid, errors = validate_generated_files(generated_files)
    if not is_valid:
        logger.error(f"Invalid files generated: {', '.join(errors)}")
        raise HTTPException(422, detail=f"Invalid files generated: {', '.join(errors)}")
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
        zipf.writestr("analysis_report.md", analysis)
        zipf.write(original_excel, f"original_{original_excel.name}")
        
        for filename, content in generated_files.items():
            zipf.writestr(f"generated/{filename}", content)
        
        # Add manifest with validation info
        manifest = {
            "generated_files": len(generated_files),
            "python_files": len([f for f in generated_files if f.endswith('.py')]),
            "validation": {
                "python_syntax_valid": all(
                    validate_python_syntax(c) 
                    for f, c in generated_files.items() 
                    if f.endswith('.py')
                ),
                "errors": errors
            },
            "timestamp": datetime.now().isoformat()
        }
        zipf.writestr("manifest.json", json.dumps(manifest, indent=2))
    
    zip_buffer.seek(0)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=project_{uuid.uuid4().hex[:8]}.zip"}
    )

@app.post("/analyze-and-generate")
async def analyze_and_generate(
    excel_file: UploadFile = File(...),
    analysis_instructions: UploadFile = File(...),
    project_prompt: str = Form(...),
    analysis_model: str = Form("deepseek/deepseek-chat-v3-0324:free"),
    generation_model: str = Form("deepseek/deepseek-chat-v3-0324:free")
):
    """
    Unified endpoint that:
    1. Analyzes an Excel file with provided instructions
    2. Generates a Python project based on the analysis and project prompt
    
    Returns a ZIP containing:
    - Analysis report
    - Generated project files
    - Original Excel file
    - Manifest with metadata
    """
    if excel_file.size > MAX_EXCEL_SIZE:
        raise HTTPException(413, "Excel file too large")
    
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        
        try:
            # Step 1: Process Excel
            excel_path = temp_path / (excel_file.filename or "uploaded.xlsx")
            with open(excel_path, "wb") as f:
                shutil.copyfileobj(excel_file.file, f)
            
            excel_data = await process_excel(excel_path)
            instructions = (await analysis_instructions.read()).decode("utf-8")
            
            # Step 2: Generate Analysis
            analysis_prompt = generate_analysis_prompt(
                excel_data["data"],
                excel_data["formulas"],
                excel_data["vba"],
                instructions
            )
            logger.info(f"Sending analysis prompt to {analysis_model}")
            analysis = await call_llm(analysis_prompt, analysis_model, 0.1)
            
            # Step 3: Generate Project
            generation_prompt = create_generation_prompt(analysis, project_prompt)
            logger.info(f"Sending generation prompt to {generation_model}")
            generation_response = await call_llm(generation_prompt, generation_model, 0.3)
            
            # Log first 500 chars of response for debugging
            logger.info(f"LLM Response (first 500 chars):\n{generation_response[:500]}")
            
            generated_files = parse_generated_files(generation_response)
            
            if not generated_files:
                # Log the full response if no files were generated
                logger.error(f"No files generated. Full response:\n{generation_response}")
                raise HTTPException(
                    422, 
                    detail="LLM failed to generate valid files. The response didn't match the expected format."
                )
            
            # Step 4: Create Output
            return await generate_zip_output(analysis, generated_files, excel_path)
            
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Processing failed: {str(e)}\n{traceback.format_exc()}")
            raise HTTPException(500, f"Processing error: {str(e)}")

@app.get("/health")
async def health_check():
    return {"status": "healthy", "version": "3.1.0"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
